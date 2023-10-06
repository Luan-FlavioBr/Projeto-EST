import pandas as pd
from funcoesDeCalculo import *
import math
import subprocess
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
import xlwings
import os

def ler_dados_quantitativos_exel(origem):
    workbook = openpyxl.load_workbook(origem, data_only=True)
    sheet = workbook.active
    letras = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    lista = list()
    numeroCelula = 1
    ultimaLetra = ''
    ultimaCelula = 0
    for letra in letras:
        celula = ''
        while True:
            celula = sheet[f'{letra}{numeroCelula}'].value
            if celula != None:
                lista.append(celula)
                sheet[f'{letra}{numeroCelula}'] = ''
                ultimaCelula = numeroCelula
                numeroCelula += 1
                ultimaLetra = letra 
            else:
                numeroCelula = 1
                break
    return lista


def realizarMedidas(lista, origemArquivo):
    ultimaCelula = 0
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Pegar a matriz de dados
    letras = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    
    for i, numero in enumerate(lista):
        sheet[f"A{i+1}"] = numero
        if numero == lista[-1]:
            ultimaCelula = len(lista)
            ultimaLetra = 'A'

    tamanho_rol_dados = ["$A$1", f"$A${ultimaCelula}"]

    lista.sort()

    indice_proxima_letra = letras.index(ultimaLetra) + 2

    sheet.column_dimensions[letras[indice_proxima_letra]].width = 27
    celula_atual = sheet[f'{letras[indice_proxima_letra]}1']
    celula_atual.alignment = Alignment(horizontal='center')

    sheet[f"{letras[indice_proxima_letra]}1"] = "Medidas de Tendência Central"
    sheet.merge_cells(f"{letras[indice_proxima_letra]}1:{letras[indice_proxima_letra + 1]}1")

    media, medianaF, moda = central(lista)
    media = round(media, 2)

    # Média, moda, mediana, 1º e 3º quartil
    titulos_medidas = ["Média","Moda", "Mediana", "1º Quartil", "3º Quartil"]
    
    ultimaC = 0
    medidas_exc = [media, moda, medianaF, f'=_xlfn.QUARTILE.EXC($A$1:${ultimaLetra}${ultimaCelula}, 1)', f'=_xlfn.QUARTILE.EXC($A$1:${ultimaLetra}${ultimaCelula}, 3)']
    if isinstance(medidas_exc[1], int) or isinstance(medidas_exc[1], float): 
        for i in range(2, 7):
            sheet[f'{letras[indice_proxima_letra]}{i}'] = titulos_medidas[i-2]
            sheet[f'{letras[indice_proxima_letra + 1]}{i}'] = medidas_exc[i-2]
            ultimaC = 7
    else:
        i = 2
        while i < 6+len(medidas_exc[1]):
            if i != 3 and i <= len(medidas_exc)+1:
                sheet[f'{letras[indice_proxima_letra]}{i}'] = titulos_medidas[i-2]
                sheet[f'{letras[indice_proxima_letra + 1]}{i}'] = medidas_exc[i-2]               
            elif i == 3:
                for mod in medidas_exc[1]:
                    sheet[f'{letras[indice_proxima_letra]}{i}'] = 'Moda'
                    sheet[f'{letras[indice_proxima_letra + 1]}{i}'] = mod
                    if i != 6:
                        i += 1 
                    else:
                        ...
            else:
                sheet[f'{letras[indice_proxima_letra]}{i}'] = titulos_medidas[i-(len(medidas_exc))]
                sheet[f'{letras[indice_proxima_letra + 1]}{i}'] = medidas_exc[i-(len(medidas_exc))] 
            i += 1
            ultimaC = i

    #Cálculo dos Intervalos de Classe
    titulos_calculo_sturges = ["Valor Máximo","Valor Mínimo","Amplitude","Qtde Linhas (Sturges)","Tamanho da Classe"]
    maximo = max(lista)
    minimo = min(lista)
    amplitude = maximo - minimo
    qtd_linhas_sturges = 1 + 3.33 * math.log(len(lista), 10)
    qtd_linhas_sturges_arredondado = round(qtd_linhas_sturges)
    tamanho_da_classe = amplitude / qtd_linhas_sturges
    tamanho_da_classe_arredondado = round(tamanho_da_classe+0.5)
    lista_calculos_sturges = [maximo, minimo, amplitude, round(qtd_linhas_sturges,2), round(tamanho_da_classe,2)]

    # Tabela Cálculo dos Intervalos de Classe
    ultimaCelula = letras[indice_proxima_letra + 1]
    indice_proxima_letra = letras.index(ultimaCelula) + 2 
    inicio_tabela = letras[letras.index(ultimaCelula) + 2] # J

    sheet.column_dimensions[letras[indice_proxima_letra]].width = 50 
    sheet.column_dimensions[letras[indice_proxima_letra+1]].width = 10

    sheet[f"{inicio_tabela}1"] = "Cálculo dos intervalos de Classe"
    celula_atual = sheet[f'{inicio_tabela}1']
    celula_atual.alignment = Alignment(horizontal='center')
    sheet.merge_cells(f"{inicio_tabela}1:{letras[indice_proxima_letra + 1]}1")

    # Inserindo valores na tabela  Cálculo dos Intervalos de Classe
    for i in range(2, len(titulos_calculo_sturges)+2):
        sheet[f"{inicio_tabela}{i}"] = titulos_calculo_sturges[i-2]
        sheet[f"{letras[letras.index(inicio_tabela)+1]}{i}"] = lista_calculos_sturges[i-2]


    # Tabela de frequencia 
    ultimaCelula = letras[letras.index(inicio_tabela) + 1] # K
    inicio_tabela_frequencia = letras[letras.index(ultimaCelula)+2] # M
    indice_proxima_letra = letras.index(ultimaCelula) + 2

    sheet.column_dimensions[letras[indice_proxima_letra]].width = 22
    sheet.column_dimensions[letras[indice_proxima_letra+1]].width = 12
    sheet.column_dimensions[letras[indice_proxima_letra+2]].width = 24
    sheet.column_dimensions[letras[indice_proxima_letra+3]].width = 7
    sheet.column_dimensions[letras[indice_proxima_letra+4]].width = 8
    sheet.column_dimensions[letras[indice_proxima_letra+5]].width = 16

    titulos_tabela_frequencia = ["Velocidade MHz", "Ponto Médio", "Fi", "Fr", "Fr(%)", "Fr acumulado(%)"]
    for i, titulo in enumerate(titulos_tabela_frequencia):
        sheet[f'{letras[indice_proxima_letra+i]}2'] = titulo
        celula_atual = sheet[f'{letras[indice_proxima_letra+i]}2']
        celula_atual.alignment = Alignment(horizontal='center')
        

    sheet[f"{inicio_tabela_frequencia}1"] = "Velocidade dos Dispositivos para um determinado tipo de CPU"
    celula_atual = sheet[f'{inicio_tabela_frequencia}1']
    celula_atual.alignment = Alignment(horizontal='center')
    sheet.merge_cells(f"{inicio_tabela_frequencia}1:{letras[indice_proxima_letra + 5]}1")

    menor_freq = minimo
    while True:
        if menor_freq % 10 == 0 or \
        menor_freq % 5 == 0: 
            break
        else:
            menor_freq -= 1

    # Arredontando o tamanho da classe para um numero multiplo de 10 ou 5
    tamanho_da_classe_ajustada = tamanho_da_classe_arredondado
    while True:
        if tamanho_da_classe_ajustada % 10 == 0 or \
        tamanho_da_classe_ajustada % 5 == 0:
            break
        else:
            tamanho_da_classe_ajustada += 1

    # Fazendo as frequências
    lista_teste = list()
    inicio_freq = menor_freq
    lista_teste.append(int(inicio_freq))
    while True:
        if inicio_freq < maximo:
            inicio_freq += tamanho_da_classe_ajustada
            lista_teste.append(int(inicio_freq))
        else:
            break

    # Inserindo a gambiarra
    lista_padrao_tabela = [" ", "-", 0, 0, 0, 0]
    for i, padrao in enumerate(lista_padrao_tabela):
        sheet[f'{letras[indice_proxima_letra+i]}3'] = padrao
        celula_atual = sheet[f'{letras[indice_proxima_letra+i]}2']
        celula_atual.alignment = Alignment(horizontal='center')
    workbook.save(f"{origemArquivo}/histograma.xlsx")

    # Inserindo os intervalos frequencias e o ponto médio
    arrumar_exel(f"{origemArquivo}/histograma.xlsx")
    celula_fi = ''
    for i in range(0, len(lista_teste)):
        celula_ponto_medio = letras[letras.index(inicio_tabela_frequencia) + 1]
        celula_fi = letras[letras.index(inicio_tabela_frequencia) + 2]
        if i < len(lista_teste) - 1:
            sheet[f'{inicio_tabela_frequencia}{4+i}'] = f"{int(lista_teste[i])} |-- {int(lista_teste[i+1])} " # Frequencias
            sheet[f'{celula_ponto_medio}{4+i}'] = (lista_teste[i] + lista_teste[i+1]) / 2 # Ponto Médio
            sheet[f'{celula_fi}{4+i}'] = f'=COUNTIFS({tamanho_rol_dados[0]}:{tamanho_rol_dados[1]},">={lista_teste[i]}", {tamanho_rol_dados[0]}:{tamanho_rol_dados[1]}, "<{lista_teste[i+1]}")' # Fi
        if i == len(lista_teste) - 1:
            ultimaCelula = len(lista_teste)+3
            sheet[f"{inicio_tabela_frequencia}{ultimaCelula}"] = "Totais"             

    sheet[f'{celula_fi}{ultimaCelula}'] = f"=SUM({celula_fi}{3}:{celula_fi}{ultimaCelula - 1})"
    celula_soma = f'{celula_fi}{ultimaCelula}'
    workbook.save(f"{origemArquivo}/histograma.xlsx")

    # Inserindo o Fr
    celula_fr = letras[letras.index(inicio_tabela_frequencia) + 3]
    for fr in range(0, len(lista_teste)):
        if fr < len(lista_teste) - 1:
            sheet[f'{celula_fr}{4+fr}'] = f'={celula_fi}{4+fr}/{celula_soma}'

    celula_soma_fr = f"{letras[letras.index(inicio_tabela_frequencia) + 3]}{ultimaCelula}"
    sheet[celula_soma_fr] = f'=SUM({celula_fr}4:{celula_fr}{ultimaCelula-1})'
    workbook.save(f"{origemArquivo}/histograma.xlsx")

    # Inserindo Fr(%)
    celula_fr_porcento = letras[letras.index(inicio_tabela_frequencia) + 4]
    for fr_p in range(0, len(lista_teste)):
        sheet[f'{celula_fr_porcento}{4+fr_p}'] = f'={celula_fi}{4+fr_p}/{celula_soma}'
        #sheet[f'{celula_fr_porcento}{4+fr_p}'].number_format = "0.00%"

    celula_soma_fr_p = f"{celula_fr_porcento}{ultimaCelula}"
    sheet[celula_soma_fr_p] = f"=SUM({celula_fr_porcento}4:{celula_fr_porcento}{ultimaCelula-1})"
    #sheet[celula_soma_fr_p].number_format = "0.00%"

    # Inserindo Fr(%) Acumulada
    celula_fr_porcento_acumulada = letras[letras.index(inicio_tabela_frequencia) + 5]
    sheet[f'{celula_fr_porcento_acumulada}4'] = f'={celula_fr_porcento}4'
    #sheet[f'{celula_fr_porcento_acumulada}4'].number_format = '0.00%'
    for fr_p_acum in range(1, len(lista_teste)):
        if fr_p_acum < len(lista_teste) - 1:
            sheet[f'{celula_fr_porcento_acumulada}{4+fr_p_acum}'] = f'=SUM({celula_fr_porcento}{4+fr_p_acum}, {celula_fr_porcento_acumulada}{3+fr_p_acum})'
            #sheet[f'{celula_fr_porcento_acumulada}{4+fr_p_acum}'].number_format = '0.00%'

    workbook.save(f"{origemArquivo}/histograma.xlsx")
    localArquivo = f"{origemArquivo}/histograma.xlsx"
    realizar_iqr_superior_inferior(localArquivo, ultimaC)
    return localArquivo

def realizar_iqr_superior_inferior(origemArquivo, ultimaC):
    arrumar_exel(origemArquivo)
    workbook = openpyxl.load_workbook(origemArquivo, data_only=True)
    sheet = workbook.active
    
    sheet[f"C{ultimaC}"] = "IQR"
    sheet[f"C{ultimaC+1}"] = "Corte Inferior"
    sheet[f"C{ultimaC+2}"] = "Corte Superior"

    quartil_1 = sheet[f"D{ultimaC-2}"].value
    quartil_2 = sheet[f"D{ultimaC-1}"].value
    sheet[f'D{ultimaC}'] = quartil_2 - quartil_1
    iqr = sheet[f'D{ultimaC}'].value
    sheet[f'D{ultimaC+1}'] = quartil_1 - 1.5 * iqr
    sheet[f'D{ultimaC+2}'] = quartil_2 + 1.5 * iqr
    workbook.save(origemArquivo)


def arrumar_exel(origemArquivo):
    # Arrumar .xlsx para resguardar a sanidade mental do Luan!
    data = openpyxl.load_workbook(origemArquivo)
    data.save(origemArquivo)
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open(origemArquivo)
    excel_book.save()
    excel_book.close()
    excel_app.quit()


# Função para fazer a table de medidas de classe nas telas
def retornar_dados_medidas_t_central(origemArquivo):
    arrumar_exel(origemArquivo)
    data = openpyxl.load_workbook(origemArquivo, data_only=True)
    sheet = data.active

    titulos_medidas_tendencia_central = list()
    medidas_tendencia_central = list()
    
    i = 2
    while True:
        celula_atual = sheet[f"C{i}"].value
        if celula_atual != None:
            titulos_medidas_tendencia_central.append(celula_atual)
            i += 1
        else:
            break

    for f in range(2, len(titulos_medidas_tendencia_central)+2):
        celula_atual = sheet[f"D{f}"].internal_value
        medidas_tendencia_central.append(celula_atual)

    lista_arrumada = list()
    for i in range(0, len(titulos_medidas_tendencia_central)):
        lista_arrumada.append([titulos_medidas_tendencia_central[i], medidas_tendencia_central[i]])

    return lista_arrumada


# Função para fazer o gráfico
def retornar_dados_intervalo_classe(origemArquivo):
    arrumar_exel(origemArquivo)
    data = openpyxl.load_workbook(origemArquivo, data_only=True)
    sheet = data.active

    titulos_intervalos_classe = list()
    valores_intervalos_classe = list()

    i = 2
    while True:
        celula_atual = sheet[f"F{i}"].value
        if celula_atual != None:
            titulos_intervalos_classe.append(celula_atual)
            i += 1
        else:
            break
    
    for valor_intervalo in range(2, len(titulos_intervalos_classe)+2):
        celula_atual = sheet[f"G{valor_intervalo}"].value
        valores_intervalos_classe.append(celula_atual)

    lista_arrumada = list()
    for i in range(0, len(titulos_intervalos_classe)):
        lista_arrumada.append([titulos_intervalos_classe[i], valores_intervalos_classe[i]])

    return lista_arrumada


# Função para retornar a lista para fazer a table de frequencia nas telas
def retornar_dados_tabela_freq(origemArquivo):
    arrumar_exel(f'{origemArquivo}/histograma.xlsx')
    data = openpyxl.load_workbook(f'{origemArquivo}/histograma.xlsx', data_only=True)
    sheet = data.active

    titulos_tabela_freq = list()
    #Pegando os titulos da tabela
    titulos_tabela_freq.append(sheet[f'I{2}'].value) 
    titulos_tabela_freq.append(sheet[f'J{2}'].value)
    titulos_tabela_freq.append(sheet[f'K{2}'].value)
    titulos_tabela_freq.append(sheet[f'L{2}'].value)
    titulos_tabela_freq.append(sheet[f'M{2}'].value)
    titulos_tabela_freq.append(sheet[f'N{2}'].value)

    coluna_de_intervalo = list()
    # Pegando a coluna de intervalo
    i = 3
    while True: 
        celula_atual = sheet[f"I{i}"].value
        if celula_atual == None:
            coluna_de_intervalo.append(" ")
        elif celula_atual == "Totais":
            coluna_de_intervalo.append(celula_atual)
            break
        else:
            coluna_de_intervalo.append(celula_atual)
        i += 1

    #Pegando coluna de ponto médio
    coluna_ponto_medio = list()
    i = 3
    while True:
        celula_atual = sheet[f"J{i}"].value
        if celula_atual != None:
            coluna_ponto_medio.append(celula_atual)
            i += 1
        else:
            coluna_ponto_medio.append(" ")
            break

    # Pegando coluna de Fi
    coluna_fi = list()
    i = 3
    while True:
        celula_atual = sheet[f"K{i}"].value
        if celula_atual != None:
            coluna_fi.append(celula_atual)
            i += 1
        else:
            break

    # Pegando a coluna de Fr
    coluna_fr = list()
    i = 3
    while True:
        celula_atual = sheet[f"M{i}"].value
        if celula_atual != None:
            coluna_fr.append(float(celula_atual))
            i += 1
        else:
            break

    # Pegando a coluna de Fr(%)
    coluna_fr_p = list()
    i = 3
    while True:
        celula_atual = sheet[f"M{i}"].value
        if celula_atual != None:
            coluna_fr_p.append(f'{celula_atual * 100}%')
            i += 1
        else:
            break

    # Pegando a coluna de Fr Acum(%)
    coluna_fr_p_acum = list()
    i = 3
    while True:
        celula_atual = sheet[f"N{i}"].value
        if celula_atual != None:
            coluna_fr_p_acum.append(f'{celula_atual * 100}%')
            i += 1
        else:
            coluna_fr_p_acum.append(" ")
            break

    return titulos_tabela_freq, coluna_de_intervalo, coluna_ponto_medio, coluna_fi, coluna_fr, coluna_fr_p, coluna_fr_p_acum


# Função para razer o table de intervalos de classe nas telas
def dado_para_tabela(origemArquivo):
    arrumar_exel(origemArquivo)
    data = openpyxl.load_workbook(origemArquivo, data_only=True)
    sheet = data.active

    lista = list()
    titulos_tabela_freq = list()
    #Pegando os titulos da tabela
    titulos_tabela_freq.append(sheet[f'I{2}'].value) 
    titulos_tabela_freq.append(sheet[f'J{2}'].value)
    titulos_tabela_freq.append(sheet[f'K{2}'].value)
    titulos_tabela_freq.append(sheet[f'L{2}'].value)
    titulos_tabela_freq.append(sheet[f'M{2}'].value)
    titulos_tabela_freq.append(sheet[f'N{2}'].value)

    coluna_de_intervalo = list()
        # Pegando a coluna de intervalo
    i = 3
    while True: 
        celula_atual = sheet[f"I{i}"].value
        if celula_atual == None:
            coluna_de_intervalo.append(" ")
        elif celula_atual == "Totais":
            coluna_de_intervalo.append(celula_atual)
            break
        else:
            coluna_de_intervalo.append(celula_atual)
        i += 1

    for valor in range(2, len(coluna_de_intervalo)+3):
        if valor == 2:
            i = sheet[f'I{valor}'].value
            if i == None:
                i = " " 
            j = sheet[f'J{valor}'].value
            if j == None:
                j = " "
            k = sheet[f'K{valor}'].value
            l = sheet[f'L{valor}'].value
            m = sheet[f'M{valor}'].value
            n = sheet[f'N{valor}'].value
            if n == None:
                n = "-"
            lista.append([i, j, k, l, m, n])
        else:
            i = sheet[f'I{valor}'].value
            if i == None:
                i = " " 
            j = sheet[f'J{valor}'].value
            if j == None:
                j = " "
            k = sheet[f'K{valor}'].value
            l = sheet[f'L{valor}'].value
            m = sheet[f'M{valor}'].value
            m = round(float(m)* 100, 2)
            n = sheet[f'N{valor}'].value
            if n == None:
                n = "-"
            elif isinstance(m, float) or isinstance(m, int):
                n = round(float(n)* 100, 2)
                n = f"{n}%"
            lista.append([i, j, k, round(l, 4), f"{m}%", n])

    return lista


def realizar_grafico(origemArquivo, gerarExel):

    # Dados fornecidos
    todasMedidas = list(retornar_dados_tabela_freq(origemArquivo))
    dados = [(todasMedidas[0][0], todasMedidas[0][-2])]

    for i in range(0, len(todasMedidas)):
        if todasMedidas[1][i] != "Totais" and todasMedidas[4][i] != 1:
            dados.append((todasMedidas[1][i], todasMedidas[4][i]))

    # Extrai os rótulos e os valores da frequência
    rotulos = [item[0] for item in dados[1:]]
    valores = [item[1] for item in dados[1:]]

    # Remove o primeiro rótulo em branco
    rotulos.pop(0)
    valores.pop(0)

    # Cria o gráfico de barras
    plt.bar(rotulos, valores, width=1.0)

    # Adiciona rótulos e título
    plt.xlabel('Intervalo de Velocidade (MHz)')
    plt.ylabel('Frequência (%)')
    plt.title('Gráfico de Barras da Frequência de Velocidade')


    for i, v in enumerate(valores):
        plt.text(i, v + 0.02, f'{v:.2%}', ha='center', va='bottom', fontsize=10)

    # Rotaciona os rótulos do eixo x para facilitar a leitura
    plt.xticks(rotation=45, ha='right')

    max_freq = max(valores)
    plt.ylim(0, max_freq + 0.1)  # Ajuste os limites conforme necessário

    y_vals = plt.gca().get_yticks()
    plt.gca().set_yticks(y_vals)  # Define a localização dos rótulos do eixo y
    plt.gca().set_yticklabels(['{:.0f}%'.format(val * 100) for val in y_vals])  # Formatação dos rótulos

    # Mostra o gráfico
    plt.tight_layout()  # Para evitar que os rótulos fiquem cortados
    plt.savefig(f'{origemArquivo}/histograma.png',format='png',dpi = 600, bbox_inches = 'tight')
    if gerarExel:
        workbook = openpyxl.load_workbook(f"{origemArquivo}/histograma.xlsx", data_only=True)
        sheet = workbook.active
        hitogramaGrafico = Image(f'{origemArquivo}/histograma.png')
        hitogramaGrafico.width = 600
        hitogramaGrafico.height = 350
        sheet.add_image(hitogramaGrafico, "C15")
        workbook.save(f'{origemArquivo}/histograma.xlsx')
        subprocess.run(['start', '', f'{origemArquivo}/histograma.xlsx'], shell=True)
        return f'{origemArquivo}/histograma.png'
    else:
        os.remove(f'{origemArquivo}/histograma.xlsx')
        return f'{origemArquivo}/histograma.png'
