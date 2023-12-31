import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import subprocess
import pandas as pd
import matplotlib.pyplot as plt
import xlwings

def pegar_dados_qualitativos_xlsx(origem):
    workbook = openpyxl.load_workbook(origem)
    sheet = workbook.active

    lista = list()
    valor_celula = ' '
    numero_celula = 1
    while True:
        celula = f'A{numero_celula}'
        valor_celula = sheet[celula].value
        if valor_celula == None:
            break
        else:
            lista.append(valor_celula)
        numero_celula += 1
    return lista


def operacoesExel(origemArquivo, localDoArquivo, lista, tipo=1, custo_de_ocorrencia=''):
    localDoArquivo = f"{origemArquivo}/analise.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if tipo == 1:
        # Formatando largura da coluna
        sheet.column_dimensions['C'].width = 22
        sheet.column_dimensions['D'].width = 17
        sheet.column_dimensions['E'].width = 11
        sheet.column_dimensions['F'].width = 18

        # Definindo uma variável com o Obj Font
        font_negrito = Font(bold=True)

        # Pegando as informações da coluna A
        tipo_de_falha = list()
        numero_de_ocorrencias = list()
        tipo_de_falha = list(set(lista))

        # Ordenando em ordem descrecente as ocorrências!
        for ocorrencia in tipo_de_falha:
            nOcorrencia = lista.count(ocorrencia)
            numero_de_ocorrencias.append([ocorrencia, nOcorrencia])
        numero_de_ocorrencias = sorted(numero_de_ocorrencias, key=lambda x: x[1], reverse=True)


        # Títulos
        sheet["C1"] = "Tipo de Falha"
        sheet['C1'].font = font_negrito
        sheet["D1"] = "Nº de Ocorrências"
        sheet['D1'].font = font_negrito
        sheet["E1"] = "Fr(%)"
        sheet['E1'].font = font_negrito
        sheet['F1'] = 'Fr Acumulada (%)'
        sheet['F1'].font = font_negrito

        soma = 0
        for i,cell in enumerate(numero_de_ocorrencias, start=1):
            celula = f'C{i+1}'
            sheet[celula] = cell[0]
            sheet[f'D{i+1}'] = cell[1]
            soma += cell[1]
            if cell == numero_de_ocorrencias[-1]:
                sheet[f'C{len(numero_de_ocorrencias)+2}'] = "Total"
                sheet[f'C{len(numero_de_ocorrencias)+2}'].font = font_negrito
                sheet[f'D{len(numero_de_ocorrencias)+2}'] = soma

                # Fazendo Fr (%)
                for i in range(1, len(numero_de_ocorrencias)+1):
                    sheet[f'E{i+1}'] = f"=D{i+1}/ D{len(numero_de_ocorrencias)+2}"
                    #sheet[f'E{i+1}'].number_format = "0.00%"
                sheet[f'E{len(numero_de_ocorrencias)+2}'] = f'=SUM(E2:E{len(numero_de_ocorrencias)+1})'
                #sheet[f'E{len(numero_de_ocorrencias)+2}'].number_format = "0.00%"

                # Fazendo Fr Acumulada (%)
                sheet['F2'] = "=E2"
                #sheet['F2'].number_format = "0.00%"
                for i in range(3, len(numero_de_ocorrencias)+2):
                    sheet[f"F{i}"] = f"=SUM(E{i},F{i-1})"
                    #sheet[f'F{i}'].number_format = "0.00%"
                sheet[f"F{len(numero_de_ocorrencias)+2}"] = "-"
    else:

        # Formatando largura da coluna
        sheet.column_dimensions['C'].width = 22
        sheet.column_dimensions['D'].width = 17
        sheet.column_dimensions['E'].width = 24
        sheet.column_dimensions['F'].width = 11
        sheet.column_dimensions['G'].width = 11
        sheet.column_dimensions['H'].width = 18

        # Definindo uma variável com o Obj Font
        font_negrito = Font(bold=True)

        # Pegando as informações da coluna A
        tipo_de_falha = list()
        numero_de_ocorrencias = list()
        totalCusto = list()
        tipo_de_falha = list(set(lista))

        # Ordenando em ordem descrecente as ocorrências!
        for i, ocorrencia in enumerate(tipo_de_falha):
            nOcorrencia = lista.count(ocorrencia)
            totalCusto.append(nOcorrencia * custo_de_ocorrencia[i])
            numero_de_ocorrencias.append([ocorrencia, nOcorrencia, custo_de_ocorrencia[i], totalCusto[i]])
        numero_de_ocorrencias = sorted(numero_de_ocorrencias, key=lambda x: x[3], reverse=True)


        # Títulos
        sheet["C1"] = "Tipo de Falha"
        sheet['C1'].font = font_negrito
        sheet["D1"] = "Nº de Ocorrências"
        sheet['D1'].font = font_negrito
        sheet["E1"] = "Impacto de custo/ocorr."
        sheet['E1'].font = font_negrito
        sheet["F1"] = "Total"
        sheet['F1'].font = font_negrito
        sheet["G1"] = "Fr(%)"
        sheet['G1'].font = font_negrito
        sheet['H1'] = 'Fr Acumulada (%)'
        sheet['H1'].font = font_negrito

        soma = 0
        for i,cell in enumerate(numero_de_ocorrencias, start=1):
            celula = f'C{i+1}'
            sheet[celula] = cell[0]
            sheet[f'D{i+1}'] = cell[1]
            sheet[f'E{i+1}'] = cell[2]
            sheet[f'F{i+1}'] = cell[3]
            soma += cell[1]
            if cell == numero_de_ocorrencias[-1]:
                sheet[f'C{len(numero_de_ocorrencias)+2}'] = "Total"
                sheet[f'C{len(numero_de_ocorrencias)+2}'].font = font_negrito
                sheet[f'D{len(numero_de_ocorrencias)+2}'] = soma
                sheet[f'E{len(numero_de_ocorrencias)+2}'] = '-'

                # Fazendo o Total(R$)
                # for i in range(1, len(numero_de_ocorrencias)+1):
                #     sheet[f'F{i+1}'] = f"=E{i+1}*D{i+1}"
                sheet[f'F{len(numero_de_ocorrencias)+2}'] = f"=SUM(F2:F{len(numero_de_ocorrencias)+1})"

                # Fazendo Fr (%)
                for i in range(1, len(numero_de_ocorrencias)+1):
                    sheet[f'G{i+1}'] = f"=F{i+1}/ F{len(numero_de_ocorrencias)+2}"
                sheet[f'G{len(numero_de_ocorrencias)+2}'] = f'=SUM(G2:G{len(numero_de_ocorrencias)+1})'

                # Fazendo Fr Acumulada (%)
                sheet['H2'] = "=G2"
                for i in range(3, len(numero_de_ocorrencias)+2):
                    sheet[f"H{i}"] = f"=SUM(G{i},H{i-1})"
                sheet[f"H{len(numero_de_ocorrencias)+2}"] = "-"

    workbook.save(localDoArquivo)
    return localDoArquivo
    


def arrumar_exel(localDoArquivo):
    # Arrumar .xlsx para resguardar a sanidade mental do Luan!
    data = openpyxl.load_workbook(localDoArquivo)
    data.save(localDoArquivo)
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open(localDoArquivo)
    excel_book.save()
    excel_book.close()
    excel_app.quit()


def aplicacaoGraficoPareto(localDoArquivo, gerarExel=False, tipo=1):
    arrumar_exel(localDoArquivo)
    localDaImagem = os.path.dirname(localDoArquivo)
    workbook = openpyxl.load_workbook(localDoArquivo, data_only = True)
    sheet = workbook.active 
    def pegarExel():
        tipoDefeito = list()
        numeroOcorrencias = list()
        frequencia = list()
        frequenciaAcumulada = list()
        totalCusto = list()

        ultimaLinha = 0
        numeroCelula = 2

        while True:
            celula = sheet[f"C{numeroCelula}"].value
            if celula != "Total":
                tipoDefeito.append(celula)
            else:
                ultimaLinha = numeroCelula
                break
            numeroCelula += 1
        
        for numeroOcorrencia in range(2, len(tipoDefeito)+2):
            celula = sheet[f'D{numeroOcorrencia}'].value
            numeroOcorrencias.append(celula)

        for fr in range(2, len(tipoDefeito)+2):
            if tipo == 1:
                celula = sheet[f'E{fr}'].value
            else:
                celula = sheet[f'G{fr}'].value
            frequencia.append(celula * 100)

        for fracum in range(2, len(tipoDefeito)+2):
            if tipo == 1:
                celula = sheet[f'F{fracum}'].value
            else:
                celula = sheet[f'H{fracum}'].value
            frequenciaAcumulada.append(celula * 100)

        if tipo != 1:
            for fracum in range(2, len(tipoDefeito)+2):
                celula = sheet[f'F{fracum}'].value
                totalCusto.append(celula)

        if tipo == 1:
            return tipoDefeito, numeroOcorrencias, frequencia, frequenciaAcumulada, ultimaLinha
        else: 
            return tipoDefeito, numeroOcorrencias, totalCusto, frequencia, frequenciaAcumulada, ultimaLinha

    if tipo == 1:
        tipo_de_falha, numeroOcorrencias, frequencia, frequeciaAcumulada, ultimaLinha = pegarExel()
        df = pd.DataFrame({'tipo defeito' : tipo_de_falha,
                    'Nº Ocor' : numeroOcorrencias, 
                    'freq (%)': frequencia,
                    'cum (%)' : frequeciaAcumulada})
    else:
        tipo_de_falha, numeroOcorrencias, totalCusto, frequencia, frequeciaAcumulada, ultimaLinha = pegarExel()
        df = pd.DataFrame({'tipo defeito' : tipo_de_falha,
                    'Nº Ocor' : numeroOcorrencias,
                    'Total Custo' : totalCusto,
                    'freq (%)': frequencia,
                    'cum (%)' : frequeciaAcumulada})

    plt.rcParams.update({'font.size': 14})
    
    fig,ax1 = plt.subplots(figsize = (12,6))

    ax1.set_title('Pareto')

    color1 = 'tomato'
    #ax1.set_xlabel('X')

    if tipo == 1:
        ax1.set_ylabel('Nº Ocorrências',color = color1)
        bars = ax1.bar(df['tipo defeito'], df['Nº Ocor'],color = color1,edgecolor = 'orange',linewidth = 2,\
                        hatch = '*')
    else:
        ax1.set_ylabel('Impacto de Custo',color = color1)
        bars = ax1.bar(df['tipo defeito'], df['Total Custo'],color = color1,edgecolor = 'orange',linewidth = 2,\
                        hatch = '*')
    #ax1.set_ylim([-10,10])
    ax1.tick_params(axis = 'y',labelcolor = color1)

    color2 = 'black'
    ax2 = ax1.twinx() # compartilhar o mesmo eixo x
    ax2.set_ylabel('%',color = color2) 

    ax2.plot(df['tipo defeito'], df['cum (%)'],color = color2,marker = 's',markersize = 8, linestyle = '-')

    ax2.tick_params(axis = 'y',labelcolor = color2)
    ax2.set_ylim([0,120])

    for tick in ax1.get_xticklabels():
        tick.set_rotation(90)

    for x, y in zip(df['tipo defeito'], df['cum (%)']):
        ax2.annotate(f'{y:.2f}%', xy=(x, y), xytext=(5, -25), textcoords='offset points',
                    ha='center', fontsize=12)

    plt.savefig(f'{localDaImagem}/graficoPareto.png',format='png',dpi = 500, bbox_inches = 'tight')
    if gerarExel:
        paretoGrafico = Image(f'{localDaImagem}/graficoPareto.png')
        paretoGrafico.width = 600
        paretoGrafico.height = 350
        sheet.add_image(paretoGrafico, f"C{ultimaLinha + 3}")
        workbook.save(f'{localDaImagem}/analise.xlsx')
        subprocess.run(['start', '', f'{localDaImagem}/analise.xlsx'], shell=True)
        return f'{localDaImagem}/graficoPareto.png'
    else:
        os.remove(localDoArquivo)
        return f'{localDaImagem}/graficoPareto.png'


def lerarquivo(localDoArquivo, tipo):
    arrumar_exel(localDoArquivo)
    workbook = openpyxl.load_workbook(localDoArquivo, data_only = True)
    sheet = workbook.active

    tipoDefeito = list()
    linhas = list()

    ultimaLinha = 0
    numeroCelula = 2

    while True:
        celula = sheet[f"C{numeroCelula}"].value
        if celula != "Total":
            tipoDefeito.append(celula)
        else:
            ultimaLinha = numeroCelula
            break
        numeroCelula += 1
    
    soma_fr = 0
    for linha in range(2, ultimaLinha+1):
        if tipo == 1:
            if linha != 1 and linha != ultimaLinha:
                c = sheet[f"C{linha}"].value
                d = sheet[f"D{linha}"].value
                e = round(float(sheet[f"E{linha}"].value) * 100, 2)
                f = round(float(sheet[f"F{linha}"].value) * 100, 2)
                linhas.append([c, d, f'{e}%', f'{f}%'])
                soma_fr += e
            else:
                c = sheet[f"C{linha}"].value
                d = sheet[f"D{linha}"].value
                e = soma_fr
                f = sheet[f"F{linha}"].value
                linhas.append([c, d, e, f])
        else:
            if linha != 1 and linha != ultimaLinha:
                c = sheet[f"C{linha}"].value
                d = sheet[f"D{linha}"].value
                e = round(float(sheet[f"E{linha}"].value), 2)
                f = round(float(sheet[f"F{linha}"].value), 2)
                g = round(float(sheet[f"G{linha}"].value) * 100, 2)
                h = round(float(sheet[f"H{linha}"].value) * 100, 2)
                linhas.append([c, d, e, f, f'{g}%', f'{h}%'])
                soma_fr += g
            else:
                c = sheet[f"C{linha}"].value
                d = sheet[f"D{linha}"].value
                e = sheet[f"E{linha}"].value
                f = round(float(sheet[f"F{linha}"].value), 2)
                g = round(soma_fr, 1)
                h = sheet[f"H{linha}"].value
                linhas.append([c, d, e, f, g, h])

    return linhas
